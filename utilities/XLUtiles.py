import openpyxl

class XLUtils:

    @staticmethod
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def readData(file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def writeData(file, sheetName, rownum, colnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(file)
